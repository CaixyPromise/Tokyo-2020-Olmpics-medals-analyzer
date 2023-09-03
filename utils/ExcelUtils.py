from openpyxl import Workbook, load_workbook
from os import name as  os_name
from models.enums import Column, ColumnName
from subprocess import Popen

# 创建一个模板
class MakeTemplate:
    def __init__(self, columns, save_path, show = True):
        self.columns = columns
        self.show = show
        self.save_path = save_path
        self.__wb = Workbook()

    def make(self):

        ws = self.__wb.active
        ws.append(self.columns)
        self.__wb.save(self.save_path)
        # 打开文件夹并高亮显示给定的文件（仅限Windows）
        if self.show and os_name == 'nt':  # Windows
            Popen(f'explorer /select,"{self.save_path}"')

class ReadTemplate:
    def __init__(self, file_path):
        self.wb = load_workbook(file_path)
        self.ws = self.wb[self.wb.sheetnames[0]]  # 默认打开第一个工作表

    def read(self, columns, response):
        column_len = len(columns)
        dataList = []
        Append = dataList.append

        # 获取第一行（列名）
        first_row = [str(cell.value).strip() for cell in self.ws[1]]

        # 检查列名是否匹配
        if tuple(first_row) != columns:
            print(first_row)
            print(columns)
            raise Exception('列名不匹配，请检查模板')

        for row in self.ws.iter_rows(min_row=2, values_only=True):  # 跳过列名
            # 清洗数据，去除每个单元格的前后空白
            cleaned_row = tuple(str(cell).strip() if cell is not None else '' for cell in row)
            if len(cleaned_row) != column_len:
                raise Exception('列数不匹配，请检查模板')
            # 检查是否所有字段都为空，如果是，则跳过该行
            if all(cell == '' for cell in cleaned_row):
                continue
            print(cleaned_row)
            Append(response(*cleaned_row))

        return dataList